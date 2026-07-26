"""토스 쉐어링크 스크래핑에 필요한 공통 기능 모음.
gui_app.py(화면 있는 버전)와 scheduled_run.py(예약 실행용)가 같이 사용함.
"""
import os
import re
import sys
import json
from pathlib import Path

_FIXED_BROWSERS_DIR = os.path.join(os.path.expanduser("~"), ".toss_link_exporter", "browsers")
os.environ.setdefault("PLAYWRIGHT_BROWSERS_PATH", _FIXED_BROWSERS_DIR)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ── 0. 브라우저(크로미움) 자동 설치 ───────────────────────────────
def ensure_chromium_installed(log=lambda m: None):
    """다른 컴퓨터에 처음 배포했을 때, 크로미움 브라우저가 없으면 자동으로 설치함."""
    try:
        from playwright.__main__ import main as playwright_main
    except Exception as e:
        log(f"자동 설치 기능을 불러오지 못했어요: {e}")
        return False

    old_argv = sys.argv
    try:
        sys.argv = ["playwright", "install", "chromium"]
        log("처음 실행이라 브라우저(크로미움)를 설치하는 중이에요. 인터넷 연결이 필요하고, 몇 분 걸릴 수 있어요...")
        try:
            playwright_main()
        except SystemExit:
            pass
        log("브라우저 설치가 끝났어요.")
        return True
    except Exception as e:
        log(f"브라우저 자동 설치 중 오류가 났어요: {e}")
        return False
    finally:
        sys.argv = old_argv


# ── 1. 상품 목록 긁어오기 ────────────────────────────────────────
def scroll_to_load_all(page, log=lambda m: None, max_idle_rounds=3, pause=1.2):
    last_count = 0
    idle_rounds = 0
    while idle_rounds < max_idle_rounds:
        page.mouse.wheel(0, 3000)
        page.wait_for_timeout(int(pause * 1000))
        cards = page.query_selector_all('[class*="ProductCard_card__"]')
        count = len(cards)
        if count == last_count:
            idle_rounds += 1
        else:
            idle_rounds = 0
            log(f"상품 {count}개 발견...")
        last_count = count
    return last_count


def extract_products(page):
    return page.evaluate(
        """
        () => {
          const cards = document.querySelectorAll('[class*="ProductCard_card__"]');
          const results = [];
          cards.forEach((card) => {
            const get = (sel) => {
              const el = card.querySelector(sel);
              return el ? el.innerText.trim() : '';
            };
            const nameBtn = card.querySelector('[class*="ProductCard_clickArea__"]');
            const name = nameBtn ? (nameBtn.getAttribute('aria-label') || '') : '';
            const discount = get('[class*="ProductThumbnailBadge_badge__"]');

            const priceBox = card.querySelector('[class*="ProductCard_price__"]');
            let price = '', priceTag = '';
            if (priceBox) {
              const spans = priceBox.querySelectorAll('span');
              if (spans[0]) price = spans[0].innerText.trim();
              if (spans[1]) priceTag = spans[1].innerText.trim();
            }

            const unitPrice = get('[class*="ProductCard_priceInfo__"] > span:last-child');
            const rating = get('[class*="ProductCard_rating__"] span:last-child');

            const badgesTop = Array.from(card.querySelectorAll('[class*="ProductCard_badgeList__"] span'))
              .map(s => s.innerText.trim()).join(', ');
            const badgesBottom = Array.from(card.querySelectorAll('[class*="ProductCard_badgeListBottom__"] span'))
              .map(s => s.innerText.trim()).join(', ');

            const commissionEl = Array.from(card.querySelectorAll('span'))
              .find(s => s.innerText.includes('수익'));
            const commission = commissionEl ? commissionEl.innerText.trim() : '';

            const imgEl = card.querySelector('[class*="ProductCard_thumbnail__"] img');
            const imageUrl = imgEl ? imgEl.src : '';

            results.push({ name, discount, price, priceTag, unitPrice, rating, badgesTop, badgesBottom, commission, imageUrl });
          });
          return results;
        }
        """
    )


def parse_discount_percent(text):
    m = re.search(r"(\d+)\s*%", text or "")
    return int(m.group(1)) if m else None


def click_and_get_link(page, idx):
    cards = page.query_selector_all('[class*="ProductCard_card__"]')
    if idx >= len(cards):
        return ""
    card = cards[idx]
    button = card.query_selector('button:has-text("링크 발급")')
    if not button:
        return ""
    try:
        button.click()
    except Exception:
        return ""
    page.wait_for_timeout(800)
    try:
        clipboard_text = page.evaluate("navigator.clipboard.readText()")
    except Exception:
        clipboard_text = ""
    m = re.search(r"https?://\S+", clipboard_text or "")
    return m.group(0) if m else ""


def scrape_one_page(page, url, min_discount, log=lambda m: None):
    """페이지 하나(카테고리 하나)를 훑어서 상품 목록(딕셔너리 리스트)을 돌려줌.
    필요하면 최소 할인율로 거르고, 각 상품의 제휴 링크까지 받아옴."""
    log(f"페이지를 여는 중: {url}")
    try:
        page.goto(url, wait_until="domcontentloaded")
    except Exception as e:
        # 로그인이 안 되어 있으면 로그인 화면으로 자동 이동되면서 이 오류가 날 수 있음(정상적인 상황)
        log(f"(참고) 페이지 이동 중 알림: {e}")
    page.wait_for_timeout(2000)
    if "login" in page.url or "signup" in page.url:
        log("로그인이 필요한 것 같아요. 브라우저 창에서 로그인해주세요.")

    count = scroll_to_load_all(page, log)
    log(f"총 {count}개 상품을 찾았어요.")

    products = extract_products(page)

    if min_discount is not None:
        before = len(products)
        products = [p for p in products if (parse_discount_percent(p.get("discount")) or 0) >= min_discount]
        log(f"할인율 {min_discount}% 이상만 남겨요: {before}개 → {len(products)}개")

    if not products:
        return []

    log("링크를 하나씩 발급받는 중... (시간이 좀 걸려요)")
    all_cards_data = extract_products(page)
    for i, prod in enumerate(products):
        idx = next((j for j, c in enumerate(all_cards_data) if c.get("name") == prod.get("name")), None)
        link = click_and_get_link(page, idx) if idx is not None else ""
        prod["link"] = link
        prod["source_url"] = url
        label = (prod.get("name") or "")[:24]
        log(f"  [{i + 1}/{len(products)}] {label} → {'완료' if link else '실패'}")

    return products


# ── 2. 중복 제거 ─────────────────────────────────────────────────
def load_seen_names(seen_path):
    p = Path(seen_path)
    if not p.exists():
        return set()
    try:
        return set(json.loads(p.read_text(encoding="utf-8")))
    except Exception:
        return set()


def save_seen_names(seen_path, names):
    Path(seen_path).write_text(json.dumps(sorted(names), ensure_ascii=False, indent=2), encoding="utf-8")


def remove_already_seen(products, seen_names):
    return [p for p in products if p.get("name") not in seen_names]


# ── 3. 할인율 높은 순 정렬 ────────────────────────────────────────
def sort_by_discount_desc(products):
    return sorted(products, key=lambda p: parse_discount_percent(p.get("discount")) or 0, reverse=True)


# ── 4. 엑셀로 저장 ───────────────────────────────────────────────
def save_excel(products, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "상품목록"
    headers = ["할인율", "상품명", "가격", "가격태그", "단위가격", "개당수익",
               "별점(리뷰수)", "배송정보", "판매자정보", "제휴링크", "상품이미지", "출처링크"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="E8F0FE")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for prod in products:
        ws.append([
            prod.get("discount", ""),
            prod.get("name", ""),
            prod.get("price", ""),
            prod.get("priceTag", ""),
            prod.get("unitPrice", ""),
            prod.get("commission", ""),
            prod.get("rating", ""),
            prod.get("badgesTop", ""),
            prod.get("badgesBottom", ""),
            prod.get("link", ""),
            prod.get("imageUrl", ""),
            prod.get("source_url", ""),
        ])
    widths = [10, 42, 12, 12, 14, 14, 14, 16, 16, 45, 55, 40]
    for col, width in zip("ABCDEFGHIJKL", widths):
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")
    wb.save(out_path)


# ── 6. 갤러리 사이트로 바로 올리기 ─────────────────────────────────
SITE_CONFIG_DEFAULT = {"site_url": "", "admin_user": "", "admin_password": ""}


def load_site_config(config_path):
    p = Path(config_path)
    if not p.exists():
        return dict(SITE_CONFIG_DEFAULT)
    try:
        cfg = json.loads(p.read_text(encoding="utf-8"))
        return {**SITE_CONFIG_DEFAULT, **cfg}
    except Exception:
        return dict(SITE_CONFIG_DEFAULT)


def save_site_config(config_path, site_url, admin_user, admin_password):
    cfg = {"site_url": site_url, "admin_user": admin_user, "admin_password": admin_password}
    Path(config_path).write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")


def upload_to_gallery(products, site_url, admin_user, admin_password, log=lambda m: None):
    """products: 스크래핑 결과 리스트. site_url: 관리자 페이지가 있는 사이트 주소
    (예: https://toss-gallery-production.up.railway.app).
    관리자 아이디/비밀번호로 인증해서 /api/products/bulk 로 한 번에 전송함."""
    try:
        import requests
    except ImportError:
        log("사이트 업로드에 필요한 라이브러리(requests)가 설치되어 있지 않아요. "
            "'pip install requests'로 설치하거나 설치하기.bat을 다시 실행해주세요.")
        return False

    if not site_url:
        log("사이트 주소가 비어있어서 사이트 업로드는 건너뛸게요.")
        return False

    url = site_url.rstrip("/") + "/api/products/bulk"
    rows = []
    for p in products:
        badges = f"{p.get('badgesTop', '')} {p.get('badgesBottom', '')}"
        rows.append({
            "name": p.get("name", ""),
            "img": p.get("imageUrl", ""),
            "link": p.get("link", ""),
            "price": p.get("price", ""),
            "discount": p.get("discount", ""),
            "priceTag": p.get("priceTag", ""),
            "unitPrice": p.get("unitPrice", ""),
            "soldOut": ("품절" in badges) or ("매진" in badges),
        })

    try:
        res = requests.post(
            url,
            json={"rows": rows},
            auth=(admin_user, admin_password),
            timeout=30,
        )
        if res.status_code == 401:
            log("사이트 업로드 실패: 관리자 아이디/비밀번호가 맞지 않아요.")
            return False
        res.raise_for_status()
        data = res.json()
        log(f"사이트에 {data.get('added', 0)}개 상품을 바로 올렸어요!")
        return True
    except Exception as e:
        log(f"사이트 업로드 중 오류가 났어요: {e}")
        return False
# ── 5. 구글 시트 업로드 (선택 사항) ───────────────────────────────
def upload_to_google_sheet(products, credentials_path, sheet_id, log=lambda m: None):
    """credentials_path: 구글 서비스 계정 키(json) 파일 경로
    sheet_id: 구글 시트 주소의 '/d/' 와 '/edit' 사이에 있는 긴 문자열"""
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        log("구글 시트 업로드에 필요한 라이브러리가 설치되어 있지 않아요. (gspread, google-auth)")
        return False

    if not Path(credentials_path).exists():
        log(f"구글 인증 파일을 찾을 수 없어요: {credentials_path}")
        return False

    try:
        scopes = ["https://www.googleapis.com/auth/spreadsheets"]
        creds = Credentials.from_service_account_file(credentials_path, scopes=scopes)
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(sheet_id)
        ws = sh.sheet1
        ws.clear()
        headers = ["할인율", "상품명", "가격", "가격태그", "단위가격", "개당수익",
                   "별점(리뷰수)", "배송정보", "판매자정보", "제휴링크", "상품이미지", "출처링크"]
        rows = [headers]
        for prod in products:
            rows.append([
                prod.get("discount", ""), prod.get("name", ""), prod.get("price", ""),
                prod.get("priceTag", ""), prod.get("unitPrice", ""), prod.get("commission", ""),
                prod.get("rating", ""), prod.get("badgesTop", ""), prod.get("badgesBottom", ""),
                prod.get("link", ""), prod.get("imageUrl", ""), prod.get("source_url", ""),
            ])
        ws.update(rows)
        log("구글 시트 업로드 완료!")
        return True
    except Exception as e:
        log(f"구글 시트 업로드 중 오류: {e}")
        return False
